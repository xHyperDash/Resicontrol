"""
Generador de reportes PDF y CSV para ResiControl.
"""

import csv
import os
from datetime import datetime

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from config import DB_PATH, REPORTES_DIR
from models import obtener_historial
from logger import logger



def generar_pdf(fecha_ini: str, fecha_fin: str, tipo: str, usuario: str) -> tuple[bool, str]:
    """
    Genera un reporte PDF de accesos entre dos fechas.
    Retorna (éxito, mensaje_o_ruta).
    """
    try:
        filas = obtener_historial_between(fecha_ini, fecha_fin, tipo)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_pdf = f"reporte_{fecha_ini}_a_{fecha_fin}_{ts}.pdf"
        ruta_pdf = os.path.join(REPORTES_DIR, nombre_pdf)
        doc = SimpleDocTemplate(
            ruta_pdf, pagesize=letter,
            leftMargin=36, rightMargin=36,
            topMargin=48, bottomMargin=36,
        )
        estilos = getSampleStyleSheet()
        elementos = []

        titulo = Paragraph(
            f"<b>ResiControl — Reporte de Accesos</b><br/>{fecha_ini} al {fecha_fin}",
            estilos["Title"],
        )
        elementos.append(titulo)
        elementos.append(Spacer(1, 16))
        elementos.append(
            Paragraph(
                f"Generado por: {usuario}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                estilos["Normal"],
            )
        )
        elementos.append(Spacer(1, 20))

        encabezados = ["Tipo", "Nombre", "Cédula", "Placa", "Entrada", "Salida", "Operador"]
        datos_tabla = [encabezados] + [
            [str(v) if v is not None else "—" for v in row] for row in filas
        ]

        tabla = Table(datos_tabla, repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("PADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elementos.append(tabla)
        elementos.append(Spacer(1, 20))
        elementos.append(Paragraph(f"Total de registros: {len(filas)}", estilos["Normal"]))

        doc.build(elementos)
        logger.info(f"PDF generado: {nombre_pdf} ({len(filas)} registros)")
        return True, ruta_pdf

    except Exception as e:
        logger.error(f"Error generando PDF: {e}")
        return False, str(e)


def generar_csv(fecha_ini: str, fecha_fin: str, tipo: str, busq: str = "") -> tuple[bool, str]:
    """
    Genera un archivo CSV de accesos entre dos fechas.
    Retorna (éxito, ruta).
    """
    try:
        filas = obtener_historial_between(fecha_ini, fecha_fin, tipo, busq)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_csv = f"reporte_{fecha_ini}_a_{fecha_fin}_{ts}.csv"
        ruta_csv = os.path.join(REPORTES_DIR, nombre_csv)
        encabezados = ["Tipo", "Nombre", "Cédula", "Placa", "Entrada", "Salida", "Operador"]

        with open(ruta_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(encabezados)
            for row in filas:
                writer.writerow([str(v) if v is not None else "" for v in row])

        logger.info(f"CSV generado: {nombre_csv} ({len(filas)} registros)")
        return True, ruta_csv

    except Exception as e:
        logger.error(f"Error generando CSV: {e}")
        return False, str(e)


def generar_xlsx(fecha_ini: str, fecha_fin: str, tipo: str, busq: str = "") -> tuple[bool, str]:
    """
    Genera un archivo XLSX (Excel) de accesos entre dos fechas.
    Retorna (éxito, ruta).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        filas = obtener_historial_between(fecha_ini, fecha_fin, tipo, busq)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_xlsx = f"reporte_{fecha_ini}_a_{fecha_fin}_{ts}.xlsx"
        ruta_xlsx = os.path.join(REPORTES_DIR, nombre_xlsx)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Accesos"

        encabezados = ["Tipo", "Nombre", "Cédula", "Placa", "Entrada", "Salida", "Operador"]
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style="thin", color="cbd5e1"),
            right=Side(style="thin", color="cbd5e1"),
            top=Side(style="thin", color="cbd5e1"),
            bottom=Side(style="thin", color="cbd5e1"),
        )

        for ci, h in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        alt_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
        for ri, row in enumerate(filas, 2):
            for ci, v in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=str(v) if v is not None else "")
                cell.font = Font(size=9)
                cell.border = thin_border
                if ri % 2 == 0:
                    cell.fill = alt_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

        wb.save(ruta_xlsx)
        logger.info(f"XLSX generado: {nombre_xlsx} ({len(filas)} registros)")
        return True, ruta_xlsx

    except ImportError:
        msg = "Librería openpyxl no instalada"
        logger.warning(msg)
        return False, msg
    except Exception as e:
        logger.error(f"Error generando XLSX: {e}")
        return False, str(e)


def obtener_historial_between(fecha_ini: str, fecha_fin: str, tipo: str, busq: str = "") -> list:
    """Obtiene historial filtrado por fechas, tipo y texto de búsqueda."""
    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    query = """
        SELECT tipo, nombre, cedula, placa, entrada, salida, operador
        FROM accesos
        WHERE DATE(entrada) BETWEEN ? AND ?
    """
    params: list = [fecha_ini, fecha_fin]
    if tipo != "Todos":
        query += " AND tipo=?"
        params.append(tipo)
    if busq:
        query += " AND (nombre LIKE ? OR cedula LIKE ? OR placa LIKE ?)"
        params += [f"%{busq}%"] * 3
    query += " ORDER BY entrada DESC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return rows